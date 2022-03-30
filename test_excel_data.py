#pytest+excel+生成日志
# 连接excel：可以在excel中增删改查测试数据
# 自动生成日志信息并将测试结果写入日志中
#pytest .\testpackage\test_topics.py -s
#注意：需要配合excel（mao.xlsx)和日志（logger.py)
# allure
# 1.//'test_excel_data'='需要运行的文件'，'./result3'='在当前路径生成运行结果'（注意路径）
# 运行：PS D:\appdata\pycharm\testpackage> pytest test_excel_data.py --alluredir ./result3
#
# 2.运行结果并自动打开allure界面
# PS D:\appdata\pycharm\testpackage> allure serve ./result3   //"./result3"="上一步生成文件”

from openpyxl import load_workbook
import pytest
import requests
import json
from testpackage.logger import logger


wb = load_workbook('D:/appdata/pycharm/testpackage/mao.xlsx')  # excel在文件中存放地址
print(wb.worksheets)
ws = wb['topics']  # topics是excel下标表名
# print(ws['A1'].value)  #输出excel中列表值

print(len(tuple(ws.rows)))
# for x in range(1,):cd
test_data = []
for x in range(2,len(tuple(ws.rows)) +1):
    testpackage_data = []
    for y in range(2,7):
        testpackage_data.append(ws.cell(row=x, column=y).value)
        print(ws.cell(row=x, column=y).value)
    test_data.append(testpackage_data)
print(test_data)  # 形成对应列表方便后续自动化测试


@pytest.mark.parametrize('url,method,topic_data,code,msg', test_data)
def test_create_topic(url, method, topic_data, code, msg):
    if method == 'post':
        res = requests.post(url, data=json.loads(topic_data))
        logger.debug(f'发送请求:{res}')
        assert res.status_code == code
        # assert res.json() == json.loads(msg)






